import csv
import io
import re
from datetime import datetime, timezone
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

from app.models import Lead


class ExportService:
    EXCEL_COLUMNS = [
        ("ID", 6),
        ("Data da Coleta", 16),
        ("Nicho Pesquisado", 20),
        ("Cidade Pesquisada", 16),
        ("Empresa", 30),
        ("Categoria", 20),
        ("Descrição", 40),
        ("Site", 30),
        ("Status do Site", 14),
        ("Telefone", 18),
        ("WhatsApp", 30),
        ("Email", 30),
        ("Instagram", 30),
        ("Facebook", 30),
        ("Endereço", 40),
        ("Bairro", 20),
        ("Cidade", 16),
        ("Estado", 8),
        ("CEP", 12),
        ("Fonte", 20),
        ("URL da Fonte", 40),
        ("Qualidade", 12),
        ("Score", 8),
        ("Tem Site?", 12),
        ("Tem Landing Page?", 16),
        ("Tem WhatsApp?", 14),
        ("Instagram Ativo?", 16),
        ("Observações Automáticas", 40),
        ("Oportunidade Identificada", 40),
        ("Serviço Sugerido", 40),
        ("Próxima Ação", 50),
        ("Contato Realizado?", 14),
        ("Data 1º Contato", 16),
        ("Data Último Contato", 16),
        ("Status Comercial", 18),
        ("Resposta do Lead", 40),
        ("Recusou Proposta?", 14),
        ("Motivo da Recusa", 40),
        ("Interessado?", 12),
        ("Follow-up Agendado?", 16),
        ("Data Follow-up", 16),
        ("Proposta Enviada?", 14),
        ("Valor Estimado", 14),
        ("Fechamento Possível?", 16),
        ("Observações Manuais", 40),
        ("Dono Identificado?", 16),
        ("Nome do Dono", 24),
        ("Falou com", 16),
        ("Canal do Contato", 16),
        ("Temperatura", 12),
        ("Tags", 24),
        ("Campanha", 20),
    ]

    def _lead_row(self, lead) -> list:
        return [
            lead.id,
            lead.collected_at.strftime("%d/%m/%Y %H:%M") if lead.collected_at else "",
            lead.niche,
            lead.city,
            lead.company_name,
            lead.category,
            lead.description,
            lead.website,
            lead.website_status,
            lead.phone,
            lead.whatsapp_link,
            lead.email,
            lead.instagram,
            lead.facebook,
            lead.address,
            lead.neighborhood,
            lead.lead_city,
            lead.state,
            lead.zipcode,
            lead.source,
            lead.source_url,
            lead.quality,
            lead.score,
            "Sim" if lead.has_website else "Não",
            "Sim" if lead.has_landing_page else "Não",
            "Sim" if lead.has_whatsapp else "Não",
            "Sim" if lead.instagram_active else "Não",
            lead.auto_notes,
            lead.opportunity,
            lead.suggested_service,
            lead.next_action,
            "Sim" if lead.contacted else "Não",
            lead.first_contact_date.strftime("%d/%m/%Y") if lead.first_contact_date else "",
            lead.last_contact_date.strftime("%d/%m/%Y") if lead.last_contact_date else "",
            lead.commercial_status,
            lead.lead_response,
            "Sim" if lead.refused else "Não",
            lead.refusal_reason,
            "Sim" if lead.interested else "Não",
            "Sim" if lead.follow_up_scheduled else "Não",
            lead.follow_up_date.strftime("%d/%m/%Y") if lead.follow_up_date else "",
            "Sim" if lead.proposal_sent else "Não",
            lead.estimated_value,
            "Sim" if lead.possible_close else "Não",
            lead.manual_notes,
            "Sim" if lead.owner_identified else "Não",
            lead.owner_name,
            lead.spoke_with,
            lead.contact_channel,
            lead.temperature,
            lead.tags,
            lead.campaign.name if lead.campaign else "",
        ]

    def export_excel(self, leads: list) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Leads"

        # Header style
        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Write headers
        for col_idx, (header_name, col_width) in enumerate(self.EXCEL_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = col_width

        # Data rows
        data_font = Font(name="Calibri", size=10)
        alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        for row_idx, lead in enumerate(leads, 2):
            row_data = self._lead_row(lead)
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=(col_idx in [6, 7]))
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

                # Conditional formatting for quality
                if col_idx == 23:  # Quality column
                    if value == "quente":
                        cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                    elif value == "morno":
                        cell.fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
                    elif value == "frio":
                        cell.fill = PatternFill(start_color="6BCB77", end_color="6BCB77", fill_type="solid")

            # Freeze panes
            ws.freeze_panes = "A2"

        # Auto-filter
        ws.auto_filter.ref = ws.dimensions

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def export_csv(self, leads: list) -> str:
        output = io.StringIO()
        writer = csv.writer(output, delimiter=";", quoting=csv.QUOTE_ALL)

        # Header
        headers = [h[0] for h in self.EXCEL_COLUMNS]
        writer.writerow(headers)

        # Data
        for lead in leads:
            row = self._lead_row(lead)
            writer.writerow(row)

        return output.getvalue()
